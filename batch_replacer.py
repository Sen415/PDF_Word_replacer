# -*- coding: utf-8 -*-
"""
批量 PDF 文字替换工具 — 核心模块，提供通用的批量替换功能，支持：

适用场景：
  - 邀请函 / 证书 / 通知批量生成：模板中署名、抬头按人员表替换
  - 合同 / 协议批量填写：将模板占位符替换为实际数据
  - 任何"一份模板 + 一张数据表 → 多份 PDF"的场景

模板语法:
  {{列名}}           → 取该列的值
  {{列名|short}}     → 取该列的值并提取短称谓（仅对职务类列名有意义）
"""

import re
import sys
import traceback
from pathlib import Path
from typing import Optional

from pdf_replacer import replace_from_list, verify

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


# ─── 职务简称映射（按优先级排列：更长、更具体的在前）───
TITLE_SHORT_NAMES = [
    "副主任医师", "主任医师", "高级工程师", "副研究员",
    "总会计师", "副部长", "副院长", "副所长", "副处长",
    "副主任", "负责人", "研究员", "秘书长", "部长",
    "院长", "书记", "教授", "会长", "处长", "科长",
    "主任", "专员", "委员", "组长",
]

# ─── 常见数据列名关键词（用于自动发现 Excel 中的有效数据列）───
COMMON_COLUMN_KEYWORDS = [
    "姓名", "人员", "名字",
    "职务", "职位", "职称", "岗位",
    "单位", "公司", "部门", "机构",
    "电话", "手机", "邮箱", "地址",
    "编号", "序号", "日期", "备注",
]


# ═══════════════════════════════════════════════════════════
#  模板字符串：{{列名}} 和 {{列名|short}}
# ═══════════════════════════════════════════════════════════

TEMPLATE_RE = re.compile(r"\{\{(.+?)\}\}")


def simplify_title(title: str) -> str:
    """将完整职务简化为称谓，例如 '设备科科长' -> '科长'"""
    cleaned = re.sub(r"\s+", "", str(title or ""))
    if not cleaned:
        return ""
    parts = [p for p in re.split(r"[、,，;；/]+", cleaned) if p]
    candidates = list(reversed(parts or [cleaned]))
    for part in candidates:
        for short_name in TITLE_SHORT_NAMES:
            if part.endswith(short_name) or short_name in part:
                return short_name
    return candidates[0]


def resolve_template(template: str, row_data: dict) -> str:
    """将模板中的 {{列名}} / {{列名|short}} 替换为当前行的实际数据。

    {{姓名}}           → 取"姓名"列的值
    {{职务|short}}     → 取"职务"列的值并提取短称谓
    """
    def _replace(match):
        expr = match.group(1).strip()
        if "|" in expr:
            col_name, *filters = [x.strip() for x in expr.split("|")]
        else:
            col_name, filters = expr, []

        # 在 row_data 中按原始列名或规范化列名查找
        value = str(row_data.get(col_name, ""))
        if not value:
            norm_col = normalize_header(col_name)
            for k, v in row_data.items():
                if normalize_header(k) == norm_col:
                    value = str(v)
                    break

        if "short" in filters:
            value = simplify_title(value)

        return value

    return TEMPLATE_RE.sub(_replace, template)


def extract_column_refs(template: str) -> list:
    """提取模板中引用的所有列名（去掉 filter 部分），去重保持顺序"""
    refs = []
    seen = set()
    for match in TEMPLATE_RE.finditer(template):
        expr = match.group(1).strip()
        col_name = expr.split("|")[0].strip()
        if col_name not in seen:
            refs.append(col_name)
            seen.add(col_name)
    return refs


# ═══════════════════════════════════════════════════════════
#  文件名工具
# ═══════════════════════════════════════════════════════════

def sanitize_filename_part(value: str) -> str:
    """去掉 Windows 文件名非法字符"""
    cleaned = re.sub(r'[<>:"/\\|?*]', "", str(value or "").strip())
    return cleaned.rstrip(". ")


def strip_salutation_punctuation(value: str) -> str:
    """去掉抬头末尾标点"""
    return re.sub(r"[\s:：,，;；.。、]+$", "", str(value or "").strip())


def derive_filename_suffix(input_pdf: str, target_text: str) -> str:
    """从模板 PDF 文件名中提取"目标文字之后"的后半段，作为输出文件名后缀。

    例如模板文件名为 "张三院长邀请函【会议通知】研讨会.pdf"：
      target_text = "张三院长："
      → 返回 "邀请函【会议通知】研讨会.pdf"

    常用于：模板 PDF 文件名已包含活动名称，只需将人名部分替换即可。
    """
    template = Path(input_pdf)
    stem = template.stem
    ext = template.suffix or ".pdf"

    target_prefix = sanitize_filename_part(strip_salutation_punctuation(target_text))
    compact_target = re.sub(r"\s+", "", target_prefix)

    # 先尝试精确匹配前缀
    for candidate in (target_prefix, compact_target):
        if not candidate:
            continue
        if stem.startswith(candidate):
            suffix = stem[len(candidate):]
            if suffix:
                return suffix + ext
        compact_stem = re.sub(r"\s+", "", stem)
        if compact_stem.startswith(candidate):
            suffix = compact_stem[len(candidate):]
            if suffix:
                return suffix + ext

    # 回退：寻找常见标记词
    for marker in ["邀请函", "通知", "证书", "合同", "协议", "报告"]:
        idx = stem.find(marker)
        if idx >= 0:
            return stem[idx:] + ext

    print("[!!] 未能从模板文件名识别后缀，将使用完整模板名: %s" % template.name)
    return "_" + template.name


# ═══════════════════════════════════════════════════════════
#  Excel 读取（通用数据表，不预设列结构）
# ═══════════════════════════════════════════════════════════

def normalize_header(value: str) -> str:
    """规范化 Excel 表头（去空格、去星号、去括号注释）"""
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[\*＊:：]", "", text)
    text = re.sub(r"[（(].*?[）)]", "", text)
    return text


def auto_detect_data_columns(headers: list) -> dict:
    """在表头行中自动发现有效数据列。

    返回 {列索引: 列名}，只保留名称包含已知关键词的列。
    """
    data_cols = {}
    for idx, header in enumerate(headers):
        name = str(header).strip() if header is not None else ""
        if not name:
            continue
        norm = normalize_header(name)
        for keyword in COMMON_COLUMN_KEYWORDS:
            kw_norm = normalize_header(keyword)
            if kw_norm in norm or norm in kw_norm:
                data_cols[idx] = name
                break
    return data_cols


def find_data_header_row(rows: list) -> tuple:
    """在前 20 行内寻找包含有效数据列的表头行。

    返回 (表头行索引, {列索引: 列名}, 表头列表)
    """
    for row_idx, row in enumerate(rows[:20]):
        headers = [str(cell).strip() if cell is not None else "" for cell in row]
        data_cols = auto_detect_data_columns(headers)
        if len(data_cols) >= 2:
            return row_idx, data_cols, headers
    raise ValueError("未找到有效的数据表头行（至少需要 2 列包含已知关键词）")


def _find_target_sheet_name(wb) -> str:
    """自动定位最可能包含数据的工作表（按关键词匹配）"""
    target_keywords = ["专家", "人员", "嘉宾", "邀请明细", "参会",
                       "数据", "名单", "信息", "花名册", "通讯录"]
    sheet_names = getattr(wb, "sheetnames", [])
    if not sheet_names:
        return wb.active.title if hasattr(wb.active, "title") else str(wb.active)
    for sheet_name in sheet_names:
        for keyword in target_keywords:
            if keyword in sheet_name:
                return sheet_name
    return sheet_names[0]


def _try_read_data_rows(ws, sheet_name: str,
                        required_cols: Optional[list] = None) -> Optional[list]:
    """尝试从工作表读取数据行；失败返回 None。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    try:
        header_row_idx, data_cols, _ = find_data_header_row(rows)
    except ValueError:
        return None

    # 若调用方指定了必需列，检查是否存在（模糊匹配）
    if required_cols:
        col_names = [normalize_header(n) for n in data_cols.values()]
        for req in required_cols:
            req_norm = normalize_header(req)
            if not any(req_norm in cn or cn in req_norm for cn in col_names):
                return None  # 缺少必要列，跳过此 sheet

    records = []
    for row_num, row in enumerate(rows[header_row_idx + 1:],
                                  start=header_row_idx + 2):
        row_data = {}
        for col_idx, col_name in data_cols.items():
            value = row[col_idx] if col_idx < len(row) else None
            row_data[col_name] = str(value or "").strip()

        # 跳过完全空行
        if not any(v for v in row_data.values()):
            continue

        row_data["_row"] = row_num
        records.append(row_data)

    print("[%s] 发现 %d 个数据列, %d 行数据" %
          (sheet_name, len(data_cols), len(records)))
    return records


def load_data_from_excel(excel_path: str,
                         sheet_name: Optional[str] = None,
                         required_cols: Optional[list] = None) -> list:
    """从 Excel 读取数据表，返回 list[dict]。

    自动识别表头行和数据列。当 sheet_name 未指定时，自动扫描所有工作表。
    required_cols 用于指定必须存在的列名（模糊匹配），用于过滤无关 sheet。
    """
    try:
        import openpyxl
    except ImportError:
        print("需要 openpyxl 库: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError("工作表 '%s' 不存在，可用: %s" %
                             (sheet_name, wb.sheetnames))
        ws = wb[sheet_name]
        result = _try_read_data_rows(ws, sheet_name, required_cols)
        wb.close()
        if result is None:
            raise ValueError("工作表 '%s' 未找到有效数据列" % sheet_name)
        return result

    # ── 自动扫描：优先匹配关键词 sheet → 逐个尝试 ──
    target_name = _find_target_sheet_name(wb)
    ws = wb[target_name]
    result = _try_read_data_rows(ws, target_name, required_cols)
    if result is not None:
        wb.close()
        return result

    for name in wb.sheetnames:
        if name == target_name:
            continue
        ws = wb[name]
        result = _try_read_data_rows(ws, name, required_cols)
        if result is not None:
            wb.close()
            return result

    wb.close()
    raise ValueError("未在任何工作表中找到有效数据列，可用 sheets: %s" %
                     wb.sheetnames)


# ═══════════════════════════════════════════════════════════
#  核心：通用批量替换
# ═══════════════════════════════════════════════════════════

def batch_replace(
        input_pdf: str,
        excel_path: str,
        output_dir: str,
        find_text: str,
        replace_template: str,
        filename_template: str = "{{姓名}}.pdf",
        sheet_name: Optional[str] = None,
        page_num: int = 0,
        verify_result: bool = False,
        overwrite: bool = False,
        dry_run: bool = False,
) -> dict:
    """批量替换 PDF 文字并生成个性化文件。

    参数
    ----
    input_pdf : str
        模板 PDF 路径。
    excel_path : str
        数据 Excel 路径。
    output_dir : str
        输出目录。
    find_text : str
        在模板 PDF 中查找的**固定文字**（不支持模板变量）。
    replace_template : str
        替换为的文字模板，支持 ``{{列名}}`` 和 ``{{列名|short}}``。
    filename_template : str
        输出文件名模板，支持 ``{{列名}}`` 和 ``{{列名|short}}``。
    sheet_name : str, optional
        Excel 工作表名称，默认自动发现。
    page_num : int
        在第几页执行替换（0-based，默认第 1 页）。
    verify_result : bool
        是否在生成后验证关键字存在。
    overwrite : bool
        是否覆盖已存在的同名文件。
    dry_run : bool
        True 时只打印预览，不实际生成文件。

    返回
    ----
    dict — 包含 total_rows / ok / fail / skipped / outputs 的统计信息。
    """
    # 提取 replace 和 filename 中引用的列名（find_text 是固定文字）
    required_cols = extract_column_refs(replace_template)
    required_cols += extract_column_refs(filename_template)
    required_cols = list(dict.fromkeys(required_cols))  # 去重保序

    data_rows = load_data_from_excel(excel_path, sheet_name,
                                     required_cols=required_cols)

    output_path = Path(output_dir)
    if not dry_run:
        output_path.mkdir(parents=True, exist_ok=True)

    stats = {
        "total_rows": len(data_rows),
        "ok": 0,
        "fail": 0,
        "skipped": 0,
        "outputs": [],
    }

    if not data_rows:
        print("[!!] 未从 Excel 读取到有效数据: %s" % excel_path)
        return stats

    print("数据列: %s" % ", ".join(
        k for k in data_rows[0].keys() if k != "_row"))
    print("查找文字: %s" % find_text)
    print("替换模板: %s" % replace_template)

    for row_data in data_rows:
        actual_replace = resolve_template(replace_template, row_data)
        filename = sanitize_filename_part(
            resolve_template(filename_template, row_data))
        if not filename.lower().endswith(".pdf"):
            filename += ".pdf"
        pdf_path = output_path / filename

        # 跳过已存在文件
        if pdf_path.exists() and not overwrite:
            stats["skipped"] += 1
            if dry_run:
                print("[DRY-RUN] 跳过（已存在）: %s" % filename)
            continue

        if dry_run:
            print("[DRY-RUN] '%s' → '%s' → %s" %
                  (find_text, actual_replace, filename))
            stats["ok"] += 1
            continue

        # 覆盖模式：先删除旧文件
        if pdf_path.exists() and overwrite:
            try:
                pdf_path.unlink()
            except OSError:
                pass

        edits = [(page_num, find_text, actual_replace, None, None, None)]

        try:
            result = replace_from_list(input_pdf, str(pdf_path), edits)
            if result["total"] > 0:
                stats["ok"] += 1
                stats["outputs"].append(str(pdf_path))
                if verify_result:
                    verify(str(pdf_path), [(page_num, actual_replace)])
            else:
                stats["fail"] += 1
                row_label = row_data.get("_row", "?")
                print("[!!] 第 %s 行: '%s' 未在 PDF 中找到" %
                      (row_label, find_text))
        except Exception:
            stats["fail"] += 1
            row_label = row_data.get("_row", "?")
            print("[!!] 第 %s 行生成失败" % row_label)
            traceback.print_exc()

    return stats
