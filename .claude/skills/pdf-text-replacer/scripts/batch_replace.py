# -*- coding: utf-8 -*-
"""
批量 PDF 文字替换工具 — 一份模板 + 一张数据表 → 多份个性化 PDF

用法:
  python batch_replace.py -i 模板.pdf -e 数据.xlsx -d 输出/ \\
      --find "尊敬的来宾：" --replace "尊敬的{{姓名}}{{职务|short}}："

模板语法:
  {{列名}}            取该列的值
  {{列名|short}}      取该列的值并提取短称谓（如 设备科科长 → 科长）
"""

import argparse
import re
import sys
import traceback
from pathlib import Path
from typing import Optional

import pymupdf
from replace_pdf_text import replace_text, verify_replacements

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# ═══════════════════════════════════════════════════════════
#  职务简称映射（按优先级排列：更长、更具体的在前）
# ═══════════════════════════════════════════════════════════

TITLE_SHORT_NAMES = [
    "副主任医师", "主任医师", "高级工程师", "副研究员",
    "总会计师", "副部长", "副院长", "副所长", "副处长",
    "副主任", "负责人", "研究员", "秘书长", "部长",
    "院长", "书记", "教授", "会长", "处长", "科长",
    "主任", "专员", "委员", "组长",
]

# ═══════════════════════════════════════════════════════════
#  常见数据列名关键词（用于自动发现 Excel 中的有效数据列）
# ═══════════════════════════════════════════════════════════

COMMON_COLUMN_KEYWORDS = [
    "姓名", "人员", "名字",
    "职务", "职位", "职称", "岗位",
    "单位", "公司", "部门", "机构",
    "电话", "手机", "邮箱", "地址",
    "编号", "序号", "日期", "备注",
]

# ═══════════════════════════════════════════════════════════
#  模板引擎：{{列名}} 和 {{列名|short}}
# ═══════════════════════════════════════════════════════════

TEMPLATE_RE = re.compile(r"\{\{(.+?)\}\}")


def simplify_title(title: str) -> str:
    """将完整职务简化为称谓，例如 '设备科科长' -> '科长'"""
    cleaned = re.sub(r"\s+", "", str(title or ""))
    if not cleaned:
        return ""
    parts = [p for p in re.split(r"[、,，;；/]+", cleaned) if p]
    candidates = list(reversed(parts or [cleaned]))
    for part in candidates:
        for short_name in TITLE_SHORT_NAMES:
            if part.endswith(short_name) or short_name in part:
                return short_name
    return candidates[0]


def resolve_template(template: str, row_data: dict) -> str:
    """将模板中的 {{列名}} / {{列名|short}} 替换为当前行的实际数据。"""
    def _replace(match):
        expr = match.group(1).strip()
        if "|" in expr:
            col_name, *filters = [x.strip() for x in expr.split("|")]
        else:
            col_name, filters = expr, []

        value = str(row_data.get(col_name, ""))
        if not value:
            norm_col = normalize_header(col_name)
            for k, v in row_data.items():
                if normalize_header(k) == norm_col:
                    value = str(v)
                    break

        if "short" in filters:
            value = simplify_title(value)

        return value

    return TEMPLATE_RE.sub(_replace, template)


def extract_column_refs(template: str) -> list:
    """提取模板中引用的所有列名（去掉 filter 部分），去重保持顺序"""
    refs = []
    seen = set()
    for match in TEMPLATE_RE.finditer(template):
        expr = match.group(1).strip()
        col_name = expr.split("|")[0].strip()
        if col_name not in seen:
            refs.append(col_name)
            seen.add(col_name)
    return refs


# ═══════════════════════════════════════════════════════════
#  文件名工具
# ═══════════════════════════════════════════════════════════

def sanitize_filename_part(value: str) -> str:
    """去掉 Windows 文件名非法字符"""
    cleaned = re.sub(r'[<>:"/\\|?*]', "", str(value or "").strip())
    return cleaned.rstrip(". ")


def strip_salutation_punctuation(value: str) -> str:
    """去掉抬头末尾标点"""
    return re.sub(r"[\s:：,，;；.。、]+$", "", str(value or "").strip())


def derive_filename_suffix(input_pdf: str, target_text: str) -> str:
    """从模板 PDF 文件名中提取"目标文字之后"的后半段，作为输出文件名后缀。

    例如模板文件名为 "张三院长邀请函【会议通知】研讨会.pdf"：
      target_text = "张三院长："
      → 返回 "邀请函【会议通知】研讨会.pdf"
    """
    template = Path(input_pdf)
    stem = template.stem
    ext = template.suffix or ".pdf"

    target_prefix = sanitize_filename_part(strip_salutation_punctuation(target_text))
    compact_target = re.sub(r"\s+", "", target_prefix)

    for candidate in (target_prefix, compact_target):
        if not candidate:
            continue
        if stem.startswith(candidate):
            suffix = stem[len(candidate):]
            if suffix:
                return suffix + ext
        compact_stem = re.sub(r"\s+", "", stem)
        if compact_stem.startswith(candidate):
            suffix = compact_stem[len(candidate):]
            if suffix:
                return suffix + ext

    for marker in ["邀请函", "通知", "证书", "合同", "协议", "报告"]:
        idx = stem.find(marker)
        if idx >= 0:
            return stem[idx:] + ext

    print("[!!] 未能从模板文件名识别后缀，将使用完整模板名: %s" % template.name)
    return "_" + template.name


# ═══════════════════════════════════════════════════════════
#  Excel 读取（通用数据表，不预设列结构）
# ═══════════════════════════════════════════════════════════

def normalize_header(value: str) -> str:
    """规范化 Excel 表头（去空格、去星号、去括号注释）"""
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[\*＊:：]", "", text)
    text = re.sub(r"[（(].*?[）)]", "", text)
    return text


def auto_detect_data_columns(headers: list) -> dict:
    """在表头行中自动发现有效数据列。返回 {列索引: 列名}"""
    data_cols = {}
    for idx, header in enumerate(headers):
        name = str(header).strip() if header is not None else ""
        if not name:
            continue
        norm = normalize_header(name)
        for keyword in COMMON_COLUMN_KEYWORDS:
            kw_norm = normalize_header(keyword)
            if kw_norm in norm or norm in kw_norm:
                data_cols[idx] = name
                break
    return data_cols


def find_data_header_row(rows: list) -> tuple:
    """在前 20 行内寻找包含有效数据列的表头行。返回 (表头行索引, {列索引: 列名}, 表头列表)"""
    for row_idx, row in enumerate(rows[:20]):
        headers = [str(cell).strip() if cell is not None else "" for cell in row]
        data_cols = auto_detect_data_columns(headers)
        if len(data_cols) >= 2:
            return row_idx, data_cols, headers
    raise ValueError("未找到有效的数据表头行（至少需要 2 列包含已知关键词）")


def _find_target_sheet_name(wb) -> str:
    """自动定位最可能包含数据的工作表（按关键词匹配）"""
    target_keywords = ["专家", "人员", "嘉宾", "邀请明细", "参会",
                       "数据", "名单", "信息", "花名册", "通讯录"]
    sheet_names = getattr(wb, "sheetnames", [])
    if not sheet_names:
        return wb.active.title if hasattr(wb.active, "title") else str(wb.active)
    for sheet_name in sheet_names:
        for keyword in target_keywords:
            if keyword in sheet_name:
                return sheet_name
    return sheet_names[0]


def _try_read_data_rows(ws, sheet_name: str,
                        required_cols: Optional[list] = None) -> Optional[list]:
    """尝试从工作表读取数据行；失败返回 None。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    try:
        header_row_idx, data_cols, _ = find_data_header_row(rows)
    except ValueError:
        return None

    if required_cols:
        col_names = [normalize_header(n) for n in data_cols.values()]
        for req in required_cols:
            req_norm = normalize_header(req)
            if not any(req_norm in cn or cn in req_norm for cn in col_names):
                return None

    records = []
    for row_num, row in enumerate(rows[header_row_idx + 1:],
                                  start=header_row_idx + 2):
        row_data = {}
        for col_idx, col_name in data_cols.items():
            value = row[col_idx] if col_idx < len(row) else None
            row_data[col_name] = str(value or "").strip()

        if not any(v for v in row_data.values()):
            continue

        row_data["_row"] = row_num
        records.append(row_data)

    print("[%s] 发现 %d 个数据列, %d 行数据" %
          (sheet_name, len(data_cols), len(records)))
    return records


def load_data_from_excel(excel_path: str,
                         sheet_name: Optional[str] = None,
                         required_cols: Optional[list] = None) -> list:
    """从 Excel 读取数据表，返回 list[dict]。

    自动识别表头行和数据列。当 sheet_name 未指定时，自动扫描所有工作表。
    """
    try:
        import openpyxl
    except ImportError:
        print("需要 openpyxl 库: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError("工作表 '%s' 不存在，可用: %s" %
                             (sheet_name, wb.sheetnames))
        ws = wb[sheet_name]
        result = _try_read_data_rows(ws, sheet_name, required_cols)
        wb.close()
        if result is None:
            raise ValueError("工作表 '%s' 未找到有效数据列" % sheet_name)
        return result

    # 自动扫描：优先匹配关键词 sheet → 逐个尝试
    target_name = _find_target_sheet_name(wb)
    ws = wb[target_name]
    result = _try_read_data_rows(ws, target_name, required_cols)
    if result is not None:
        wb.close()
        return result

    for name in wb.sheetnames:
        if name == target_name:
            continue
        ws = wb[name]
        result = _try_read_data_rows(ws, name, required_cols)
        if result is not None:
            wb.close()
            return result

    wb.close()
    raise ValueError("未在任何工作表中找到有效数据列，可用 sheets: %s" %
                     wb.sheetnames)


# ═══════════════════════════════════════════════════════════
#  核心：通用批量替换
# ═══════════════════════════════════════════════════════════

def batch_replace(
        input_pdf: str,
        excel_path: str,
        output_dir: str,
        find_text: str,
        replace_template: str,
        filename_template: str = "{{姓名}}.pdf",
        sheet_name: Optional[str] = None,
        page_num: int = 0,
        verify_result: bool = False,
        overwrite: bool = False,
        dry_run: bool = False,
) -> dict:
    """批量替换 PDF 文字并生成个性化文件。

    参数
    ----
    input_pdf : 模板 PDF 路径。
    excel_path : 数据 Excel 路径。
    output_dir : 输出目录。
    find_text : 在模板 PDF 中查找的**固定文字**（不支持模板变量）。
    replace_template : 替换为的文字模板，支持 ``{{列名}}`` 和 ``{{列名|short}}``。
    filename_template : 输出文件名模板。
    sheet_name : Excel 工作表名称，默认自动发现。
    page_num : 在第几页执行替换（0-based，默认第 1 页）。
    verify_result : 是否在生成后验证关键字存在。
    overwrite : 是否覆盖已存在的同名文件。
    dry_run : True 时只打印预览，不实际生成文件。

    返回: dict — 包含 total_rows / ok / fail / skipped / outputs 的统计信息。
    """
    required_cols = extract_column_refs(replace_template)
    required_cols += extract_column_refs(filename_template)
    required_cols = list(dict.fromkeys(required_cols))

    data_rows = load_data_from_excel(excel_path, sheet_name,
                                     required_cols=required_cols)

    output_path = Path(output_dir)
    if not dry_run:
        output_path.mkdir(parents=True, exist_ok=True)

    stats = {
        "total_rows": len(data_rows),
        "ok": 0,
        "fail": 0,
        "skipped": 0,
        "outputs": [],
    }

    if not data_rows:
        print("[!!] 未从 Excel 读取到有效数据: %s" % excel_path)
        return stats

    print("数据列: %s" % ", ".join(
        k for k in data_rows[0].keys() if k != "_row"))
    print("查找文字: %s" % find_text)
    print("替换模板: %s" % replace_template)

    for row_data in data_rows:
        actual_replace = resolve_template(replace_template, row_data)
        filename = sanitize_filename_part(
            resolve_template(filename_template, row_data))
        if not filename.lower().endswith(".pdf"):
            filename += ".pdf"
        pdf_path = output_path / filename

        if pdf_path.exists() and not overwrite:
            stats["skipped"] += 1
            if dry_run:
                print("[DRY-RUN] 跳过（已存在）: %s" % filename)
            continue

        if dry_run:
            print("[DRY-RUN] '%s' → '%s' → %s" %
                  (find_text, actual_replace, filename))
            stats["ok"] += 1
            continue

        if pdf_path.exists() and overwrite:
            try:
                pdf_path.unlink()
            except OSError:
                pass

        try:
            doc = pymupdf.open(input_pdf)
            n = replace_text(doc, page_num, find_text, actual_replace)
            if n > 0:
                doc.subset_fonts()
                doc.save(str(pdf_path))
                doc.close()
                stats["ok"] += 1
                stats["outputs"].append(str(pdf_path))
                if verify_result:
                    verify_replacements(str(pdf_path), [{
                        "page_num": page_num,
                        "replace": actual_replace,
                    }])
            else:
                doc.close()
                stats["fail"] += 1
                row_label = row_data.get("_row", "?")
                print("[!!] 第 %s 行: '%s' 未在 PDF 中找到" %
                      (row_label, find_text))
        except Exception:
            stats["fail"] += 1
            row_label = row_data.get("_row", "?")
            print("[!!] 第 %s 行生成失败" % row_label)
            traceback.print_exc()

    return stats


# ═══════════════════════════════════════════════════════════
#  命令行入口
# ═══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="批量 PDF 文字替换工具 — 一份模板 + 一张数据表 → 多份个性化 PDF",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
模板语法:
  {{列名}}          取该列的值
  {{列名|short}}    取该列的值并提取短称谓（如 设备科科长 → 科长）

示例:
  python batch_replace.py -i 模板.pdf -e 人员表.xlsx -d 输出/ \\
      --find "尊敬的来宾：" --replace "尊敬的{{姓名}}{{职务|short}}："
  python batch_replace.py -i 模板.pdf -e 数据.xlsx -d 输出/ \\
      --find "XXX" --replace "{{姓名}}" --dry-run
        """,
    )

    parser.add_argument("--input", "-i", required=True,
                        help="模板 PDF 路径")
    parser.add_argument("--excel", "-e", required=True,
                        help="数据 Excel 路径")
    parser.add_argument("--output-dir", "-d", required=True,
                        help="输出目录")
    parser.add_argument("--find", "-f", required=True,
                        help="查找的固定文字")
    parser.add_argument("--replace", "-r", required=True,
                        help="替换模板，支持 {{列名}} / {{列名|short}}")
    parser.add_argument("--filename", "-n", default=None,
                        help="输出文件名模板。留空则自动: {{姓名}}.pdf")
    parser.add_argument("--filename-suffix-from", metavar="PDF",
                        default=None,
                        help="从指定 PDF 文件名推导输出文件名后缀")
    parser.add_argument("--sheet", "-s", default=None,
                        help="Excel 工作表名（留空则自动发现）")
    parser.add_argument("--page", "-p", type=int, default=1,
                        help="替换目标页码，从 1 开始（默认: 1）")
    parser.add_argument("--verify", "-v", action="store_true",
                        help="生成后验证替换文字是否存在")
    parser.add_argument("--overwrite", action="store_true",
                        help="覆盖已存在的同名文件")
    parser.add_argument("--dry-run", action="store_true",
                        help="预览模式：只显示将要执行的操作，不实际生成文件")

    args = parser.parse_args()

    page_0based = args.page - 1

    filename_template = args.filename
    if not filename_template and args.filename_suffix_from:
        suffix = derive_filename_suffix(args.filename_suffix_from, args.find)
        filename_template = "{{姓名}}{{职务|short}}" + suffix
        print("从模板文件名推导后缀: %s" % suffix)
    elif not filename_template:
        filename_template = "{{姓名}}.pdf"

    print("=" * 50)
    print("模板 PDF : %s" % args.input)
    print("数据 Excel: %s" % args.excel)
    print("输出目录  : %s" % args.output_dir)
    print("查找文字  : %s" % args.find)
    print("替换模板  : %s" % args.replace)
    print("文件名    : %s" % filename_template)
    print("目标页码  : 第 %d 页" % args.page)
    if args.dry_run:
        print("*** 预览模式（不生成文件）***")
    print("=" * 50)

    stats = batch_replace(
        input_pdf=args.input,
        excel_path=args.excel,
        output_dir=args.output_dir,
        find_text=args.find,
        replace_template=args.replace,
        filename_template=filename_template,
        sheet_name=args.sheet,
        page_num=page_0based,
        verify_result=args.verify,
        overwrite=args.overwrite,
        dry_run=args.dry_run,
    )

    print("=" * 50)
    if args.dry_run:
        print("预览完成 — 共 %d 行数据，将生成 %d 份 PDF" %
              (stats["total_rows"], stats["ok"]))
    else:
        print("完成 — 数据 %d 行, 生成 %d 份, 跳过 %d 份, 失败 %d 份" %
              (stats["total_rows"], stats["ok"], stats["skipped"], stats["fail"]))
        if stats["outputs"]:
            print("保存目录: %s" % args.output_dir)


if __name__ == "__main__":
    main()
